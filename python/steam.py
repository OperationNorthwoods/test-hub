import requests
import json
from config import API_KEY, STEAM_ID, EMAIL, FILENAME
from datetime import datetime
import openpyxl
from openpyxl.styles.borders import Border, Side
import ezgmail

# setting up seperate date and time variables
current_datetime = datetime.now()
current_date = current_datetime.strftime('%Y-%m-%d')
current_time = current_datetime.strftime('%H:%M:%S')
# api call url
url = f'https://api.steampowered.com/IPlayerService/GetRecentlyPlayedGames/v0001/?key={API_KEY}&steamid={STEAM_ID}&format=json '
response = requests.get(url)
# Setup the Excel sheet
wb = openpyxl.load_workbook(f'{FILENAME}')
sheet = wb.active
max = sheet.max_row
max_Iter = max + 1  # make sure we don't go too far down rows
col_A, col_B, col_C, col_D, col_E, col_F = range(1, 7)
# All Columns should have an equal max_row
#
# {FILENAME} Column Cheatsheet
# A=Date B=Time C=Game D=TotalHours
# E=HoursAddedSinceLastRan F=Last2WeeksHours

if response.status_code == 200:  # 200 = it worked

    # Parse the JSON data
    data = json.loads(response.content)

    gameCounter = 0  # counts number of games in json response
    gameDict = {}  # storing games with game as key, and minutes as values in a list

    for game in data['response']['games']:
        gameCounter += 1
        name = game['name']
        playtime_forever = game['playtime_forever']
        playtime_2weeks = game['playtime_2weeks']
        gameDict[name] = (playtime_forever, playtime_2weeks)
    
    # Start writing to workbook
    temp = list(gameDict.keys())  # makes a list of all keys in dict aka every game's name
    tempCount = 0  # we use this as index of 'temp' so we only access 1 game name at a time

    # Main writing loop; loops throught all columns in one row, one row at a time.
    while max_Iter <= max+gameCounter:
        sheet.cell(row=max_Iter, column=col_A, value=current_date)
        sheet.cell(row=max_Iter, column=col_B, value=current_time)
        sheet.cell(row=max_Iter, column=col_C, value=temp[tempCount])
        sheet.cell(row=max_Iter, column=col_D, value=gameDict[temp[tempCount]][0])
        sheet.cell(row=max_Iter, column=col_F, value=gameDict[temp[tempCount]][1])

        # logic to determine value of column_E
        if max != 1:
            current_game = temp[tempCount]
            new_hours = gameDict[temp[tempCount]][0]
            found_match = False
            for i in range(max, 0, -1):
                if sheet.cell(row=i, column=col_C).value == current_game:
                    previous_hours = sheet.cell(row=i, column=col_D).value
                    if isinstance(previous_hours, (int, float)):
                        hours_since_last_played = new_hours - previous_hours
                    sheet.cell(row=max_Iter, column=col_E, value=hours_since_last_played)
                    found_match = True
                    break
            # if not found_match:                 # I dont think this works
            #     hours_since_last_played = 'N/A'
            #     sheet.cell(row=max_Iter, column=col_E, value=hours_since_last_played)

        # adds border below last row
        if max_Iter == max+gameCounter:
            border = Border(bottom=Side(style='thick'))
            for i in range(col_A, col_F+1):
                sheet.cell(row=max_Iter, column=i).border = border
    

        tempCount += 1
        max_Iter += 1

    print('Sucess. Data Fetched from API and written to workbook.')
    ezgmail.send(f'{EMAIL}', 'Sucess!', f'Data Fetched from API and written to workbook. {current_datetime}.')
else:
    sheet.cell(row=max_Iter, column=col_A, value=current_date)
    sheet.cell(row=max_Iter, column=col_B, value=current_time)
    ezgmail.send(f'{EMAIL}', 'Error!', f'Error fetching data from API. {current_datetime}.')
    print("Error fetching data from API.")

wb.save(f'{FILENAME}')

